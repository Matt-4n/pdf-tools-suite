import os
import re
import fitz  # PyMuPDF
import csv
import argparse
import sys
import logging
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from pdf_optimizer import PDFOptimizer

# For Excel file reading
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    try:
        import xlrd
        EXCEL_AVAILABLE = True
    except ImportError:
        EXCEL_AVAILABLE = False

def setup_logging(job_id=None):
    """Set up simple logging for PDF merger"""
    
    # Create logs directory if it doesn't exist
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    # Create log filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if job_id:
        log_filename = f"merger_{job_id}_{timestamp}.log"
    else:
        log_filename = f"pdf_merger_{timestamp}.log"
    
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            # Log to file
            logging.FileHandler(log_dir / log_filename, encoding='utf-8'),
            # Also show on screen (console)
            logging.StreamHandler()
        ]
    )
    
    # Return logger
    logger = logging.getLogger('PDFMerger')
    logger.info("=== PDF Merger Started ===")
    logger.info(f"Log file: {log_filename}")
    
    return logger

class PDFMerger:
    def __init__(self, input_folder, output_folder, reference_doc=None, edi_file=None, 
                 enable_optimization=True, target_size_mb=1.2):
        # Add logging
        self.logger = logging.getLogger('PDFMerger')
        self.logger.info(f"Initializing PDF Merger: {input_folder} -> {output_folder}")
        
        # Basic initialization
        self.input_folder = Path(input_folder)
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        self.clients = defaultdict(lambda: {'info': None, 'pages': []})
        self.manifest = {}
        self.reference_doc = reference_doc
        self.edi_file = edi_file
        
        # Add optimization settings
        self.enable_optimization = enable_optimization
        self.target_size_mb = target_size_mb
        
        if self.enable_optimization:
            self.optimizer = PDFOptimizer(target_size_mb=target_size_mb, quality=85)
            self.logger.info(f"PDF optimization enabled (target: {target_size_mb}MB)")
            
            # Initialize optimization statistics
            self.optimization_stats = {
                'files_optimized': 0,
                'total_savings_mb': 0,
                'average_compression_ratio': 0
            }
        
        # Priority order for creating manifest:
        # 1. EDI Excel file (most reliable)
        # 2. Reference PDF document
        # 3. Existing CSV manifest
        if edi_file:
            self.create_manifest_from_edi(edi_file)
        elif reference_doc:
            self.create_manifest_from_reference(reference_doc)
        else:
            # Try to load existing CSV manifest
            self.load_manifest("client_manifest.csv")
        
        # NEW: Add tracking for reports
        self.compression_report = []
        self.tax_alerts = [] 

    def create_manifest_from_edi(self, edi_file_path):
        """Create CSV manifest from EDI Excel file"""
        edi_path = Path(edi_file_path)
        if not edi_path.exists():
            self.logger.error(f"EDI file not found: {edi_file_path}")
            return

        if not EXCEL_AVAILABLE:
            self.logger.error("Excel libraries not available. Please install openpyxl or xlrd.")
            return

        self.logger.info(f"Creating manifest from EDI file: {edi_file_path}")

        try:
            manifest_data = {}

            # Check file extension and use appropriate library
            if str(edi_path).endswith('.xls'):
                # Use xlrd for .xls files
                import xlrd
                workbook = xlrd.open_workbook(edi_path)
                sheet = workbook.sheet_by_index(0)

                # Read data starting from row 1 (skip header row 0)
                for row_idx in range(1, sheet.nrows):
                    row = sheet.row_values(row_idx)
                    if len(row) > 11:  # Ensure we have enough columns
                        consignee_name = row[6]  # Column 6: "Consignees Name"
                        consignee_ref = row[11]  # Column 11: "Consignees Reference"

                        if consignee_ref and consignee_name:
                            # Clean up the reference format
                            ref_clean = str(consignee_ref).strip()
                            name_clean = str(consignee_name).strip()

                            # For EDI files, trust the data completely - no validation
                            if ref_clean and name_clean and ref_clean != 'nan' and name_clean != 'nan':
                                manifest_data[ref_clean] = name_clean
                                self.logger.debug(f"Added EDI entry: {ref_clean} -> {name_clean}")

            else:
                # Use openpyxl for .xlsx files
                from openpyxl import load_workbook
                workbook = load_workbook(edi_path)
                sheet = workbook.active

                # Read data starting from row 2 (skip header row 1)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > 11:  # Ensure we have enough columns
                        consignee_name = row[6]  # Column 6: "Consignees Name"
                        consignee_ref = row[11]  # Column 11: "Consignees Reference"

                        if consignee_ref and consignee_name:
                            # Clean up the reference format
                            ref_clean = str(consignee_ref).strip()
                            name_clean = str(consignee_name).strip()

                            # For EDI files, trust the data completely - no validation
                            if ref_clean and name_clean and ref_clean != 'nan' and name_clean != 'nan':
                                manifest_data[ref_clean] = name_clean
                                self.logger.debug(f"Added EDI entry: {ref_clean} -> {name_clean}")

            # Create CSV file
            csv_path = "client_manifest.csv"
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['ConsigneeRef', 'FullName'])

                for ref, name in sorted(manifest_data.items()):
                    writer.writerow([ref, name])

            self.logger.info(f"Created manifest CSV with {len(manifest_data)} clients: {csv_path}")

            # Load the created manifest into memory
            self.manifest = manifest_data

        except Exception as e:
            self.logger.error(f"Error creating manifest from EDI file: {e}")

    def create_manifest_from_reference(self, reference_doc):
        """Create manifest from reference PDF document - placeholder implementation"""
        self.logger.info(f"Creating manifest from reference document: {reference_doc}")
        # TODO: Implement this method if needed
        pass

    def load_manifest(self, manifest_file):
        """Load existing CSV manifest"""
        self.logger.info(f"Loading manifest from file: {manifest_file}")
        try:
            import csv
            with open(manifest_file, 'r', newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    ref = row.get('ConsigneeRef', '').strip()
                    name = row.get('FullName', '').strip()
                    if ref and name:
                        self.manifest[ref] = name
                        self.logger.info(f"Loaded client: {ref} -> {name}")
            
            self.logger.info(f"Successfully loaded {len(self.manifest)} clients from manifest")
            
        except Exception as e:
            self.logger.error(f"Failed to load manifest: {e}")

    def scan_for_tax_keywords(self, doc, client_name, client_ref):
        """
        Scan every page after page 12 of a document for tax-relevant keywords
        """
        tax_keywords = ['tools', 'alcohol', 'new']
        found_keywords = []

        total_pages = len(doc)
        # Scan all pages after page 12 (manifest section)
        start_page = 12  # Start from page 13 (0-indexed = 12)

        if total_pages <= 12:
            self.logger.debug(f"Document only has {total_pages} pages, skipping tax scan: {client_name}")
            return

        self.logger.debug(f"Scanning pages {start_page + 1}-{total_pages} for tax keywords: {client_name}")

        for page_num in range(start_page, total_pages):
            try:
                page = doc[page_num]
                text = page.get_text().lower()

                for keyword in tax_keywords:
                    if keyword in text:
                        # Find context around the keyword
                        words = text.split()
                        for i, word in enumerate(words):
                            if keyword in word:
                                # Get 3 words before and after for context
                                context_start = max(0, i - 3)
                                context_end = min(len(words), i + 4)
                                context = ' '.join(words[context_start:context_end])

                                found_keywords.append({
                                    'keyword': keyword.upper(),
                                    'page': page_num + 1,
                                    'context': context.strip()
                                })
                                self.logger.info(f"🚨 TAX ALERT: {keyword.upper()} found in {client_name} (Page {page_num + 1})")
                                break  # Only record first occurrence per page

            except Exception as e:
                self.logger.debug(f"Could not scan page {page_num + 1}: {e}")

        # Store tax alerts for this client
        if found_keywords:
            self.tax_alerts.append({
                'client_name': client_name,
                'client_ref': client_ref,
                'alerts': found_keywords
            })

    def generate_compression_report(self):
        """
        Generate a detailed compression report
        """
        if not self.compression_report:
            return "No compression data available."

        report = []
        report.append("=" * 80)
        report.append("📊 PDF COMPRESSION REPORT")
        report.append("=" * 80)

        total_original = sum(item['original_size_mb'] for item in self.compression_report)
        total_final = sum(item['final_size_mb'] for item in self.compression_report)
        total_saved = total_original - total_final

        report.append(f"📈 SUMMARY:")
        report.append(f"   • Files processed: {len(self.compression_report)}")
        report.append(f"   • Total space saved: {total_saved:.2f} MB")
        report.append(f"   • Average compression: {total_original/total_final:.1f}x")
        report.append(f"   • Overall size reduction: {(total_saved/total_original)*100:.1f}%")
        report.append("")

        report.append("📁 INDIVIDUAL FILES:")
        report.append("-" * 80)

        # Sort by savings (largest first)
        sorted_files = sorted(self.compression_report, key=lambda x: x['savings_mb'], reverse=True)

        for item in sorted_files:
            savings_pct = (item['savings_mb'] / item['original_size_mb']) * 100
            report.append(f"• {item['filename']}")
            report.append(f"  {item['original_size_mb']:.2f}MB → {item['final_size_mb']:.2f}MB "
                         f"(saved {item['savings_mb']:.2f}MB, {savings_pct:.1f}%)")
            report.append("")

        return "\n".join(report)

    def generate_tax_alert_report(self):
        """
        Generate a tax alert report for keywords found
        """
        if not self.tax_alerts:
            return "✅ No tax-relevant keywords found in any documents."

        report = []
        report.append("=" * 80)
        report.append("🚨 TAX ALERT REPORT")
        report.append("=" * 80)

        report.append(f"⚠️  FLAGGED CLIENTS: {len(self.tax_alerts)}")
        report.append(f"🔍 Keywords scanned: TOOLS, ALCOHOL, NEW")
        report.append(f"📄 Scan area: All pages after page 12 (manifest section)")
        report.append("")

        for alert in self.tax_alerts:
            report.append(f"🚨 {alert['client_name']} ({alert['client_ref']})")
            report.append("-" * 60)

            for keyword_alert in alert['alerts']:
                report.append(f"   • {keyword_alert['keyword']} found on page {keyword_alert['page']}")
                report.append(f"     Context: \"{keyword_alert['context']}\"")
            report.append("")

        return "\n".join(report)

    def save_reports_to_file(self, output_folder):
        """
        Save both reports to text files
        """
        try:
            # Compression Report
            compression_report_path = Path(output_folder) / "compression_report.txt"
            with open(compression_report_path, 'w', encoding='utf-8') as f:
                f.write(self.generate_compression_report())
            self.logger.info(f"📊 Compression report saved: {compression_report_path}")

            # Tax Alert Report
            tax_report_path = Path(output_folder) / "tax_alert_report.txt"
            with open(tax_report_path, 'w', encoding='utf-8') as f:
                f.write(self.generate_tax_alert_report())
            self.logger.info(f"🚨 Tax alert report saved: {tax_report_path}")

            return compression_report_path, tax_report_path

        except Exception as e:
            self.logger.error(f"Failed to save reports: {e}")
            return None, None

    def process_multi_client_document(self, pdf_path, doc_type):
        """Process multi-client documents (Advice of Arrivals, Bills of Lading)"""
        self.logger.info(f"Processing {doc_type}: {pdf_path.name}")
        
        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            
            # Process each page to find client references
            for page_num in range(total_pages):
                page = doc[page_num]
                text = page.get_text()
                
                # Look for reference patterns in the text
                ref_patterns = [
                    r'(\d{3}[-/]\d{3}[-/]\d{3})',  # 000-000-000 or 000/000/000
                    r'(\d{3}\s+\d{3}\s+\d{3})',   # 000 000 000
                ]
                
                for pattern in ref_patterns:
                    matches = re.findall(pattern, text)
                    for match in matches:
                        # Normalize reference format
                        ref_normalized = match.replace('-', '/').replace(' ', '/')
                        
                        # ENHANCED: Match on first 11 characters to handle suffix variations
                        ref_base = ref_normalized[:11]  # Get first 11 characters: "000/527/962"

                        # Find matching EDI reference by comparing first 11 characters
                        matched_ref = None
                        for edi_ref in self.manifest.keys():
                            if edi_ref[:11] == ref_base:
                                matched_ref = edi_ref
                                break

                        if matched_ref:
                            client_name = self.manifest[matched_ref]
                            self.logger.info(f"{doc_type}: Found {ref_normalized} -> {matched_ref} - {client_name} on page {page_num + 1}")
                        
                        # Store page info for this client
                        if not self.clients[ref_normalized]['info']:
                            self.clients[ref_normalized]['info'] = (ref_normalized, client_name)
                        
                        self.clients[ref_normalized]['pages'].append({
                            'page_num': page_num,
                            'doc_type': doc_type,
                            'doc_obj': doc
                        })
                        break  # Only match once per page
            
            return doc
            
        except Exception as e:
            self.logger.error(f"Error processing {doc_type} {pdf_path.name}: {e}")
            return None

    def process_customer_document_edi_first(self, pdf_path):
        """Process individual customer document using EDI-first matching"""
        self.logger.info(f"Processing Customer Document: {pdf_path.name}")

        # First, try to extract reference from filename
        filename = pdf_path.name
        ref_match = re.search(r'(\d{3}[-/]\d{3}[-/]\d{3})', filename)
        
        try:
            doc = fitz.open(pdf_path)
            
            if ref_match:
                file_ref = ref_match.group(1).replace('-', '/')
                
                # Check if this reference exists in EDI manifest
                if file_ref in self.manifest:
                    client_name = self.manifest[file_ref]
                    self.logger.info(f"Customer Doc: {file_ref} - {client_name} (EDI match)")
                    
                    # Store customer document info
                    if not self.clients[file_ref]['info']:
                        self.clients[file_ref]['info'] = (file_ref, client_name)
                    
                    # Add all pages of this document
                    for page_num in range(len(doc)):
                        self.clients[file_ref]['pages'].append({
                            'page_num': page_num,
                            'doc_type': 'Customer Document',
                            'doc_obj': doc
                        })
                    
                    return doc
            
            # If filename didn't work, scan document content for EDI references
            text = ""
            for page in doc:
                text += page.get_text()
            
            # Look for any EDI reference in the document text
            for edi_ref in self.manifest.keys():
                if edi_ref in text or edi_ref.replace('/', '-') in text:
                    client_name = self.manifest[edi_ref]
                    self.logger.info(f"Customer Doc: {edi_ref} - {client_name} (EDI content match)")
                    
                    if not self.clients[edi_ref]['info']:
                        self.clients[edi_ref]['info'] = (edi_ref, client_name)
                    
                    for page_num in range(len(doc)):
                        self.clients[edi_ref]['pages'].append({
                            'page_num': page_num,
                            'doc_type': 'Customer Document',
                            'doc_obj': doc
                        })
                    
                    return doc
            
            # Document doesn't match any EDI reference
            self.logger.warning(f"Customer document {pdf_path.name} does not match any EDI reference - skipping")
            doc.close()
            return None
            
        except Exception as e:
            self.logger.error(f"Error processing customer document {pdf_path.name}: {e}")
            return None

    def merge_client_documents(self, client_key, client_data):
        """Merge all documents for a specific client with optimization and tax scanning"""
        consignee_ref, full_name = client_data['info']
        if not consignee_ref:
            self.logger.error(f"Missing consignee reference for {client_key}")
            return False

        # If no name, create a default one
        if not full_name:
            full_name = f"Client_{consignee_ref.replace('/', '_')}"

        self.logger.info(f"Merging documents for: {consignee_ref} - {full_name}")

        # Group pages by document type
        advice_pages = []
        bill_pages = []
        customer_pages = []

        for page_info in client_data['pages']:
            if page_info['doc_type'] == 'Advice of Arrivals':
                advice_pages.append(page_info)
            elif page_info['doc_type'] == 'Bill of Lading':
                bill_pages.append(page_info)
            elif page_info['doc_type'] == 'Customer Document':
                customer_pages.append(page_info)

        self.logger.info(f"Found: {len(advice_pages)} Advice pages, {len(bill_pages)} Bill pages, {len(customer_pages)} Customer pages")

        # Create merged PDF in correct order: Advice → Bills → Customer
        merged_doc = fitz.open()  # Create new empty document

        # Add pages in order
        for page_info in advice_pages:
            source_doc = page_info['doc_obj']
            page_num = page_info['page_num']
            merged_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)

        for page_info in bill_pages:
            source_doc = page_info['doc_obj']
            page_num = page_info['page_num']
            merged_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)

        for page_info in customer_pages:
            source_doc = page_info['doc_obj']
            page_num = page_info['page_num']
            merged_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)

        # Save merged document
        safe_ref = consignee_ref.replace('/', '_')
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', full_name)
        output_filename = f"{safe_ref}_{safe_name}.pdf"
        output_path = self.output_folder / output_filename

        # NEW: Scan for tax keywords before saving
        self.scan_for_tax_keywords(merged_doc, full_name, consignee_ref)

        try:
            # Save the merged document
            merged_doc.save(output_path)
            original_size = output_path.stat().st_size / (1024 * 1024)  # MB
            
            # Optimize if enabled
            if self.enable_optimization:
                self.logger.info(f"Optimizing merged document: {output_filename}")
                
                # FIXED: Get the optimization result as a dictionary
                optimization_result = self.optimizer.optimize_pdf(str(output_path), str(output_path))
                
                if optimization_result and optimization_result.get('optimized'):
                    # Extract the final size from the result dictionary
                    final_size = optimization_result['final_size_mb']
                    savings = optimization_result['savings_mb']
                    
                    # Update optimization stats
                    self.optimization_stats['files_optimized'] += 1
                    self.optimization_stats['total_savings_mb'] += savings
                    
                    # Store compression data for reporting
                    self.compression_report.append({
                        'filename': output_filename,
                        'original_size_mb': optimization_result['original_size_mb'],
                        'final_size_mb': final_size,
                        'savings_mb': savings
                    })
                    
                    self.logger.info(f"✅ Saved merged document: {output_filename} ({final_size:.2f}MB, saved {savings:.2f}MB)")
                
                elif optimization_result:
                    # File didn't need optimization (already small enough)
                    final_size = optimization_result['final_size_mb']
                    self.logger.info(f"✅ Saved merged document: {output_filename} ({final_size:.2f}MB, no optimization needed)")
                    
                else:
                    # Optimization failed
                    self.logger.warning(f"Optimization failed for {output_filename}, keeping original")
                    final_size = original_size
            else:
                final_size = original_size
                self.logger.info(f"✅ Saved merged document: {output_filename} ({final_size:.2f}MB)")
            
            merged_doc.close()
            return True

        except Exception as e:
            self.logger.error(f"Error saving merged document for {consignee_ref}: {e}")
            merged_doc.close()
            return False

    def process_all_documents(self):
        """
        Enhanced process: analyze all PDFs and merge by EDI client list ONLY
        FIXED: Prevents Advice of Arrival duplication by proper file categorization
        """
        pdf_files = list(self.input_folder.glob("*.pdf"))
        
        if not pdf_files:
            self.logger.error("No PDF files found!")
            return
                
        # CRITICAL: Only proceed if we have an EDI manifest
        if not self.manifest:
            self.logger.error("No EDI manifest loaded! Cannot process without client list.")
            return
        
        self.logger.info(f"Found {len(pdf_files)} PDF files to process")
        self.logger.info(f"EDI manifest contains {len(self.manifest)} clients")
        
        # ENHANCED: Categorize files first to prevent cross-contamination
        advice_files = []
        bill_files = []
        customer_files = []
        
        self.logger.info("📂 Categorizing PDF files...")
        
        # REPLACE this section in your pdf_merger.py (around lines 815-835):

        for pdf_path in pdf_files:
            filename = pdf_path.name  # Use original case, not .lower()
            filename_lower = filename.lower()
            
            # 1. Advice of Arrival: "Advice of Arrival ICR1032499.pdf"
            if filename_lower.startswith('advice of arrival'):
                advice_files.append(pdf_path)
                self.logger.info(f"   📋 Advice of Arrival: {pdf_path.name}")
            
            # 2. Bill of Lading: "000-534-000_HBL.pdf" (ends with _HBL.pdf)
            elif filename.endswith('_HBL.pdf'):
                bill_files.append(pdf_path)
                self.logger.info(f"   🚢 Bill of Lading (HBL): {pdf_path.name}")
            
            # 3. Customer Document: "000-534-055_Document.pdf" (ends with _Document.pdf)
            elif filename.endswith('_Document.pdf'):
                customer_files.append(pdf_path)
                self.logger.info(f"   👥 Customer Document: {pdf_path.name}")
            
            # 4. Fallback for other files with client references
            elif re.search(r'\d{3}[-/]\d{3}[-/]\d{3}', filename):
                customer_files.append(pdf_path)
                self.logger.info(f"   👥 Customer Document (by reference): {pdf_path.name}")
            
            # 5. Everything else
            else:
                customer_files.append(pdf_path)
                self.logger.warning(f"   ❓ Unknown file type, treating as customer document: {pdf_path.name}")
                
        # Summary of categorization
        self.logger.info(f"📊 File categorization complete:")
        self.logger.info(f"   📋 Advice files: {len(advice_files)}")
        self.logger.info(f"   🚢 Bill files: {len(bill_files)}")
        self.logger.info(f"   👥 Customer files: {len(customer_files)}")
        
        # Validation: Multiple advice files warning
        if len(advice_files) > 1:
            self.logger.warning(f"⚠️  Multiple Advice of Arrival files detected: {[f.name for f in advice_files]}")
            self.logger.warning(f"⚠️  Will use only the FIRST file: {advice_files[0].name}")
        
        opened_docs = []  # Keep track of opened documents
        
        # FIXED: Process Advice of Arrival files (ONLY THE FIRST ONE)
        if advice_files:
            advice_file = advice_files[0]  # Take only the first one to prevent duplication
            self.logger.info(f"🔄 Processing single Advice of Arrival: {advice_file.name}")
            doc = self.process_multi_client_document(advice_file, 'Advice of Arrivals')
            if doc:
                opened_docs.append(doc)
            
            # Skip any additional advice files
            if len(advice_files) > 1:
                for skipped_file in advice_files[1:]:
                    self.logger.warning(f"⏭️  Skipping duplicate Advice file: {skipped_file.name}")
        else:
            self.logger.warning("⚠️  No Advice of Arrival files found!")
        
        # Process all Bill of Lading files
        self.logger.info(f"🔄 Processing {len(bill_files)} Bill of Lading files...")
        for bill_file in bill_files:
            self.logger.info(f"   🚢 Processing: {bill_file.name}")
            doc = self.process_multi_client_document(bill_file, 'Bill of Lading')
            if doc:
                opened_docs.append(doc)
        
        # Process customer documents
        self.logger.info(f"🔄 Processing {len(customer_files)} customer documents...")
        for customer_file in customer_files:
            self.logger.info(f"   👥 Processing: {customer_file.name}")
            doc = self.process_customer_document_edi_first(customer_file)
            if doc:
                opened_docs.append(doc)
        
        # ENHANCED: Validation before merging
        self.logger.info("🔍 Validating client data before merging...")
        
        total_clients_with_pages = 0
        for edi_ref, edi_name in self.manifest.items():
            if edi_ref in self.clients and self.clients[edi_ref]['pages']:
                # Count pages by type to detect duplication
                advice_count = len([p for p in self.clients[edi_ref]['pages'] if p['doc_type'] == 'Advice of Arrivals'])
                bill_count = len([p for p in self.clients[edi_ref]['pages'] if p['doc_type'] == 'Bill of Lading'])
                customer_count = len([p for p in self.clients[edi_ref]['pages'] if p['doc_type'] == 'Customer Document'])
                
                if advice_count > 1:
                    self.logger.warning(f"⚠️  Client {edi_ref} has {advice_count} Advice pages (expected 1)")
                
                self.logger.info(f"   ✅ {edi_ref}: {advice_count} advice, {bill_count} bill, {customer_count} customer pages")
                total_clients_with_pages += 1
        
        self.logger.info(f"📊 Validation complete: {total_clients_with_pages} clients have documents")
        
        # ONLY merge documents for EDI clients
        edi_clients_processed = 0
        
        self.logger.info(f"🔄 MERGING DOCUMENTS FOR {len(self.manifest)} EDI CLIENTS")
        
        for edi_ref, edi_name in self.manifest.items():
            # Check if this EDI client has any documents
            if edi_ref in self.clients and self.clients[edi_ref]['pages']:
                self.logger.info(f"🔧 Processing EDI client: {edi_ref} - {edi_name}")
                
                # Ensure we use the EDI name as authoritative
                self.clients[edi_ref]['info'] = (edi_ref, edi_name)
                
                success = self.merge_client_documents(edi_ref, self.clients[edi_ref])
                if success:
                    edi_clients_processed += 1
            else:
                self.logger.warning(f"⚠️  No documents found for EDI client: {edi_ref} - {edi_name}")
        
        # Close all opened documents
        self.logger.info("🔒 Closing all opened documents...")
        for doc in opened_docs:
            doc.close()
        
        # Update average compression ratio
        if self.enable_optimization and self.optimization_stats['files_optimized'] > 0:
            total_original = sum(item['original_size_mb'] for item in self.compression_report)
            total_final = sum(item['final_size_mb'] for item in self.compression_report)
            self.optimization_stats['average_compression_ratio'] = total_original / total_final if total_final > 0 else 1
        
        # Generate and save reports
        self.logger.info("📊 Generating reports...")
        
        # Print reports to console
        print("\n" + self.generate_compression_report())
        print("\n" + self.generate_tax_alert_report())
        
        # Save reports to files
        compression_file, tax_file = self.save_reports_to_file(self.output_folder)
        
        # Final summary
        self.logger.info(f"✅ Processing complete!")
        self.logger.info(f"   📁 Output folder: {self.output_folder}")
        self.logger.info(f"   👥 Clients processed: {edi_clients_processed}/{len(self.manifest)}")
        self.logger.info(f"   📄 Files processed: {len(opened_docs)} documents")
        
        if len(advice_files) > 1:
            self.logger.info(f"   ⚠️  Note: {len(advice_files)-1} duplicate Advice files were skipped")

    # ALSO ADD THIS ENHANCED LOGGING TO merge_client_documents METHOD:

    def merge_client_documents(self, client_key, client_data):
        """Merge all documents for a specific client with enhanced validation"""
        consignee_ref, full_name = client_data['info']
        if not consignee_ref:
            self.logger.error(f"Missing consignee reference for {client_key}")
            return False

        # If no name, create a default one
        if not full_name:
            full_name = f"Client_{consignee_ref.replace('/', '_')}"

        self.logger.info(f"🔧 Merging documents for: {consignee_ref} - {full_name}")

        # Group pages by document type
        advice_pages = []
        bill_pages = []
        customer_pages = []

        for page_info in client_data['pages']:
            if page_info['doc_type'] == 'Advice of Arrivals':
                advice_pages.append(page_info)
            elif page_info['doc_type'] == 'Bill of Lading':
                bill_pages.append(page_info)
            elif page_info['doc_type'] == 'Customer Document':
                customer_pages.append(page_info)

        # ENHANCED: Detailed logging and validation
        self.logger.info(f"   📊 Page counts: {len(advice_pages)} advice, {len(bill_pages)} bill, {len(customer_pages)} customer")
        
        # VALIDATION: Alert if unexpected page counts
        if len(advice_pages) > 1:
            self.logger.warning(f"   ⚠️  UNEXPECTED: {len(advice_pages)} Advice pages for client {consignee_ref} (expected 1)")
        elif len(advice_pages) == 0:
            self.logger.warning(f"   ⚠️  No Advice pages found for client {consignee_ref}")

        # Create merged PDF in correct order: Advice → Bills → Customer
        merged_doc = fitz.open()  # Create new empty document

        # Add pages in order with logging
        self.logger.info(f"   📄 Adding {len(advice_pages)} advice pages...")
        for page_info in advice_pages:
            source_doc = page_info['doc_obj']
            page_num = page_info['page_num']
            merged_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)

        self.logger.info(f"   📄 Adding {len(bill_pages)} bill pages...")
        for page_info in bill_pages:
            source_doc = page_info['doc_obj']
            page_num = page_info['page_num']
            merged_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)

        self.logger.info(f"   📄 Adding {len(customer_pages)} customer pages...")
        for page_info in customer_pages:
            source_doc = page_info['doc_obj']
            page_num = page_info['page_num']
            merged_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)

        # Continue with the rest of the existing merge_client_documents logic...
        # (Save merged document, optimization, etc. - keep existing code)
        
        # Save merged document
        safe_ref = consignee_ref.replace('/', '_')
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', full_name)
        output_filename = f"{safe_ref}_{safe_name}.pdf"
        output_path = self.output_folder / output_filename

        # Scan for tax keywords before saving
        self.scan_for_tax_keywords(merged_doc, full_name, consignee_ref)

        try:
            # Save the merged document
            merged_doc.save(output_path)
            original_size = output_path.stat().st_size / (1024 * 1024)  # MB
            
            # Optimize if enabled
            if self.enable_optimization:
                self.logger.info(f"   🗜️  Optimizing: {output_filename}")
                
                # Get the optimization result as a dictionary
                optimization_result = self.optimizer.optimize_pdf(str(output_path), str(output_path))
                
                if optimization_result and optimization_result.get('optimized'):
                    # Extract the final size from the result dictionary
                    final_size = optimization_result['final_size_mb']
                    savings = optimization_result['savings_mb']
                    
                    # Update optimization stats
                    self.optimization_stats['files_optimized'] += 1
                    self.optimization_stats['total_savings_mb'] += savings
                    
                    # Store compression data for reporting
                    self.compression_report.append({
                        'filename': output_filename,
                        'original_size_mb': optimization_result['original_size_mb'],
                        'final_size_mb': final_size,
                        'savings_mb': savings
                    })
                    
                    self.logger.info(f"   ✅ Saved: {output_filename} ({final_size:.2f}MB, saved {savings:.2f}MB)")
                
                elif optimization_result:
                    # File didn't need optimization (already small enough)
                    final_size = optimization_result['final_size_mb']
                    self.logger.info(f"   ✅ Saved: {output_filename} ({final_size:.2f}MB, no optimization needed)")
                    
                else:
                    # Optimization failed
                    self.logger.warning(f"   ⚠️  Optimization failed for {output_filename}, keeping original")
                    final_size = original_size
            else:
                final_size = original_size
                self.logger.info(f"   ✅ Saved: {output_filename} ({final_size:.2f}MB)")
            
            merged_doc.close()
            return True

        except Exception as e:
            self.logger.error(f"   ❌ Error saving merged document for {consignee_ref}: {e}")
            merged_doc.close()
            return False

def parse_command_line():
    """Parse command line arguments with optimization options"""
    parser = argparse.ArgumentParser(description='PDF Merger - Merge shipping documents by client')
    
    # Required arguments
    parser.add_argument('--input-folder', required=True,
                    help='Folder containing PDF files to merge')
    parser.add_argument('--output-folder', required=True,
                    help='Folder to save merged PDF files')
    
    # Optional arguments for manifest
    parser.add_argument('--edi-file',
                    help='EDI Excel file (.xls or .xlsx) for client manifest')
    parser.add_argument('--reference-doc',
                    help='Reference PDF document to extract client manifest')
    parser.add_argument('--manifest-file',
                    help='Existing CSV manifest file')
    
    # Optimization options
    parser.add_argument('--enable-optimization', action='store_true', default=True,
                    help='Enable PDF optimization (default: enabled)')
    parser.add_argument('--disable-optimization', action='store_true',
                    help='Disable PDF optimization')
    parser.add_argument('--target-size', type=float, default=1.2,
                    help='Target file size in MB (default: 1.2)')
    parser.add_argument('--quality', type=int, default=85,
                    help='Image quality 0-100 (default: 85)')
    
    # Optional settings
    parser.add_argument('--job-id',
                    help='Job ID for web application processing')
    parser.add_argument('--json-output', action='store_true',
                    help='Output results as JSON (for web application)')
    
    return parser.parse_args()


if __name__ == "__main__":
    # Parse command line arguments
    args = parse_command_line()
    
    # Set up logging FIRST
    logger = setup_logging(args.job_id)
    
    # Check if input folder exists
    if not Path(args.input_folder).exists():
        logger.error(f"Input folder does not exist: {args.input_folder}")
        print(f"❌ Error: Input folder does not exist: {args.input_folder}")
        sys.exit(1)
    
    # Determine optimization settings
    enable_optimization = args.enable_optimization and not args.disable_optimization
    
    # Create PDF merger with command line arguments
    try:
        logger.info("Creating PDF merger instance")
        pdf_merger = PDFMerger(
            input_folder=args.input_folder,
            output_folder=args.output_folder,
            edi_file=args.edi_file,
            reference_doc=args.reference_doc,
            enable_optimization=enable_optimization,
            target_size_mb=args.target_size
        )
        
        # Load manifest file if specified
        if args.manifest_file:
            logger.info(f"Loading manifest file: {args.manifest_file}")
            pdf_merger.load_manifest(args.manifest_file)
        
        # Process documents
        logger.info("Starting document processing")
        pdf_merger.process_all_documents()
        
        # Get final statistics
        total_clients = len(pdf_merger.clients)
        
        # Include optimization stats in output
        result_stats = {
            'processed_files': len(list(Path(args.input_folder).glob("*.pdf"))),
            'merged_clients': total_clients,
            'optimization': pdf_merger.optimization_stats if enable_optimization else None
        }
        
        # Simple success message
        if args.json_output:
            import json
            result = {
                'success': True,
                'message': 'Processing completed successfully',
                'output_folder': args.output_folder,
                'stats': result_stats,
                'tax_alerts': pdf_merger.tax_alerts  # ADD THIS LINE
            }
            print(json.dumps(result))
        else:
            print(f"\n✅ Processing complete! Check the '{args.output_folder}' folder for merged PDFs.")
            
            if enable_optimization:
                opt_stats = pdf_merger.optimization_stats
                print(f"🗜️ Optimization Summary:")
                print(f"   Files optimized: {opt_stats['files_optimized']}")
                print(f"   Total space saved: {opt_stats['total_savings_mb']:.2f} MB")
                if opt_stats['files_optimized'] > 0:
                    print(f"   Average compression: {opt_stats['average_compression_ratio']:.2f}x")
        
        logger.info("=== PDF Merger Completed Successfully ===")
    
    except Exception as e:
        error_msg = f"Processing failed: {str(e)}"
        logger.error(error_msg)
        logger.error("=== PDF Merger Failed ===")
        
        if args.json_output:
            import json
            print(json.dumps({'success': False, 'error': error_msg}))
        else:
            print(f"❌ {error_msg}")
        sys.exit(1)
